import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def read_xlsm_sheets(file_path):
    """ Reads an .xlsm file and returns the names of all sheets in it. """
    workbook = openpyxl.load_workbook(file_path, keep_vba=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    return sheet_names

def choose_sheets(sheet_names):
    """ Prompts the user to choose two sheets from the provided list. """
    for i, name in enumerate(sheet_names, 1):
        print(f"{i}. {name}")

    choices = []
    while len(choices) < 2:
        try:
            choice = int(input("Choose a sheet number (1, 2, 3, etc.): "))
            if 1 <= choice <= len(sheet_names):
                choices.append(sheet_names[choice - 1])
            else:
                print("Invalid choice. Please choose a valid sheet number.")
        except ValueError:
            print("Invalid input. Please enter a number.")

        if len(choices) == 1:
            print("Choose one more sheet.")
    
    return choices[0], choices[1]


def create_comparison_sheet(file_path, sheet1, sheet2):
    """ Creates a comparison sheet in the .xlsm file with specified columns from the chosen sheets and their differences. """
    # Define the specific columns to search for
    specific_columns = [
        "tag_pri_2axles_auto",
        "tag_pri_2axles_truck",
        "tag_pri_3axles_truck",
        "tag_pri_5axles_truck",
        "tag_pri_7axles_truck",
        "tag_pri_2axles_motorcycle"
    ]

    # Read the data from the selected sheets
    df1 = pd.read_excel(file_path, sheet_name=sheet1, engine='openpyxl')
    df2 = pd.read_excel(file_path, sheet_name=sheet2, engine='openpyxl')

    # Select the first three columns and the specific columns if they exist
    df1_selected = df1.iloc[:, :3].join(df1[specific_columns].dropna(axis=1, how='all'))
    df2_selected = df2.iloc[:, :3].join(df2[specific_columns].dropna(axis=1, how='all'))

    # Calculate the differences for the specific columns
    df_diff = df1_selected[specific_columns].subtract(df2_selected[specific_columns], fill_value=0)

    # Load the workbook with openpyxl for editing
    workbook = openpyxl.load_workbook(file_path, keep_vba=True)
    
    # Add or clear the comparison sheet
    if "Comparison Sheet" in workbook.sheetnames:
        comparison_sheet = workbook["Comparison Sheet"]
        comparison_sheet.delete_rows(1, comparison_sheet.max_row)
    else:
        comparison_sheet = workbook.create_sheet("Comparison Sheet")
    
    # Write data from df1_selected to the comparison sheet
    for r_idx, row in enumerate(dataframe_to_rows(df1_selected, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            comparison_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Leave a gap of four columns and write data from df2_selected
    offset_df2 = len(df1_selected.columns) + 2
    for r_idx, row in enumerate(dataframe_to_rows(df2_selected, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            comparison_sheet.cell(row=r_idx, column=c_idx + offset_df2, value=value)

    # Leave a gap of four columns and write the differences
    offset_diff = offset_df2 + len(df2_selected.columns) + 2
    for r_idx, row in enumerate(dataframe_to_rows(df_diff, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            comparison_sheet.cell(row=r_idx, column=c_idx + offset_diff, value=value)
    
    # Save the workbook
    workbook.save(file_path)
    workbook.close()
                    

# Execution 
file_path = "/Users/ravishankartiwari/Library/CloudStorage/GoogleDrive-ravistiwari@mapup.ai/Shared drives/data-europe/turkey/02-toll-data/O-3/00-archive/comparision-testing/O-3-testing.xlsm"  # Replace with your actual file path
sheet_names = read_xlsm_sheets(file_path)
sheet1, sheet2 = choose_sheets(sheet_names)
create_comparison_sheet(file_path, sheet1, sheet2)
